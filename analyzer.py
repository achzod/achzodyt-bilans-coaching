"""
Module d'analyse IA des bilans de coaching - Claude Sonnet 4.5
"""

import os
import base64
import json
import io
import re
import requests
from typing import List, Dict, Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from anthropic import Anthropic
from dotenv import load_dotenv
from PIL import Image

# Excel support
try:
    import openpyxl
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Warning: openpyxl not installed, Excel parsing disabled")

load_dotenv()

# Client Claude
anthropic_client = Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")

MAX_IMAGE_SIZE = 4 * 1024 * 1024


def extract_google_doc_links(text: str) -> List[str]:
    """Extrait tous les liens Google Docs d'un texte"""
    if not text:
        return []
    # Pattern pour Google Docs: docs.google.com/document/d/ID
    pattern = r'https?://docs\.google\.com/document/d/([a-zA-Z0-9_-]+)'
    matches = re.findall(pattern, text)
    return list(set(matches))  # Unique IDs


def extract_google_sheet_links(text: str) -> List[str]:
    """Extrait tous les liens Google Sheets d'un texte"""
    if not text:
        return []
    # Pattern pour Google Sheets: docs.google.com/spreadsheets/d/ID
    pattern = r'https?://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)'
    matches = re.findall(pattern, text)
    return list(set(matches))


def fetch_google_doc_content(doc_id: str) -> str:
    """
    Recupere le contenu d'un Google Doc via export public ou API
    """
    print(f"[GDOC] Fetching doc {doc_id[:20]}...")

    # Methode 1: Export public en txt (si le doc est partage "anyone with link")
    export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
    try:
        response = requests.get(export_url, timeout=15)
        if response.status_code == 200 and len(response.text) > 50:
            print(f"[GDOC] OK via export public ({len(response.text)} chars)")
            return response.text[:10000]  # Limiter a 10k chars
    except Exception as e:
        print(f"[GDOC] Export public failed: {e}")

    # Methode 2: API Google Drive avec API key
    if GOOGLE_API_KEY:
        try:
            api_url = f"https://www.googleapis.com/drive/v3/files/{doc_id}/export?mimeType=text/plain&key={GOOGLE_API_KEY}"
            response = requests.get(api_url, timeout=15)
            if response.status_code == 200:
                print(f"[GDOC] OK via Drive API ({len(response.text)} chars)")
                return response.text[:10000]
        except Exception as e:
            print(f"[GDOC] Drive API failed: {e}")

    # Methode 3: Essayer de scraper la page HTML publique
    try:
        html_url = f"https://docs.google.com/document/d/{doc_id}/pub"
        response = requests.get(html_url, timeout=15)
        if response.status_code == 200:
            # Extraire le texte du HTML
            html = response.text
            # Nettoyer le HTML
            text = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r'<[^>]+>', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            if len(text) > 100:
                print(f"[GDOC] OK via HTML pub ({len(text)} chars)")
                return text[:10000]
    except Exception as e:
        print(f"[GDOC] HTML pub failed: {e}")

    return f"[Google Doc {doc_id[:10]}... - contenu non accessible]"


def fetch_google_sheet_content(sheet_id: str) -> str:
    """
    Recupere le contenu d'un Google Sheet via export CSV public
    """
    print(f"[GSHEET] Fetching sheet {sheet_id[:20]}...")

    # Export en CSV (si partage public)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
    try:
        response = requests.get(export_url, timeout=15)
        if response.status_code == 200 and len(response.text) > 20:
            # Convertir CSV en texte lisible
            lines = response.text.strip().split('\n')
            formatted = []
            for line in lines[:50]:  # Max 50 lignes
                cells = line.split(',')
                formatted.append(" | ".join(cells))
            content = "\n".join(formatted)
            print(f"[GSHEET] OK via export CSV ({len(content)} chars)")
            return content[:8000]
    except Exception as e:
        print(f"[GSHEET] Export CSV failed: {e}")

    # Methode 2: API Google Sheets
    if GOOGLE_API_KEY:
        try:
            api_url = f"https://sheets.googleapis.com/v4/spreadsheets/{sheet_id}/values/A1:Z100?key={GOOGLE_API_KEY}"
            response = requests.get(api_url, timeout=15)
            if response.status_code == 200:
                data = response.json()
                values = data.get("values", [])
                formatted = []
                for row in values[:50]:
                    formatted.append(" | ".join(str(cell) for cell in row))
                content = "\n".join(formatted)
                print(f"[GSHEET] OK via Sheets API ({len(content)} chars)")
                return content[:8000]
        except Exception as e:
            print(f"[GSHEET] Sheets API failed: {e}")

    return f"[Google Sheet {sheet_id[:10]}... - contenu non accessible]"


def parse_all_google_docs(text: str) -> str:
    """Detecte et parse tous les Google Docs ET Sheets dans un texte"""
    contents = []

    # Google Docs
    doc_ids = extract_google_doc_links(text)
    if doc_ids:
        print(f"[GDOC] Found {len(doc_ids)} Google Doc(s) to parse")
        for doc_id in doc_ids[:3]:  # Max 3 docs
            content = fetch_google_doc_content(doc_id)
            if content and not content.startswith("[Google Doc"):
                contents.append(f"\n=== GOOGLE DOC (ID: {doc_id[:10]}...) ===\n{content}")

    # Google Sheets
    sheet_ids = extract_google_sheet_links(text)
    if sheet_ids:
        print(f"[GSHEET] Found {len(sheet_ids)} Google Sheet(s) to parse")
        for sheet_id in sheet_ids[:2]:  # Max 2 sheets
            content = fetch_google_sheet_content(sheet_id)
            if content and not content.startswith("[Google Sheet"):
                contents.append(f"\n=== GOOGLE SHEET (ID: {sheet_id[:10]}...) ===\n{content}")

    return "\n".join(contents)


def compress_image_if_needed(b64_data: str, media_type: str) -> tuple:
    raw_size = len(base64.b64decode(b64_data))
    if raw_size <= MAX_IMAGE_SIZE:
        return b64_data, media_type
    try:
        img_bytes = base64.b64decode(b64_data)
        img = Image.open(io.BytesIO(img_bytes))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        quality = 85
        max_dimension = 2000
        while True:
            if max(img.size) > max_dimension:
                ratio = max_dimension / max(img.size)
                new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
                img_resized = img.resize(new_size, Image.LANCZOS)
            else:
                img_resized = img
            buffer = io.BytesIO()
            img_resized.save(buffer, format='JPEG', quality=quality, optimize=True)
            compressed_data = buffer.getvalue()
            if len(compressed_data) <= MAX_IMAGE_SIZE:
                return base64.b64encode(compressed_data).decode('utf-8'), 'image/jpeg'
            quality -= 10
            max_dimension -= 200
            if quality < 30 or max_dimension < 800:
                img_resized = img.resize((800, int(800 * img.size[1] / img.size[0])), Image.LANCZOS)
                buffer = io.BytesIO()
                img_resized.save(buffer, format='JPEG', quality=30, optimize=True)
                return base64.b64encode(buffer.getvalue()).decode('utf-8'), 'image/jpeg'
    except Exception as e:
        print(f"Erreur compression: {e}")
        return b64_data, media_type


def detect_image_type(b64_data: str) -> Optional[str]:
    try:
        raw = base64.b64decode(b64_data[:32])
        if raw[0] == 0xFF and raw[1] == 0xD8:
            return 'image/jpeg'
        if raw[0] == 0x89 and raw[1] == 0x50:
            return 'image/png'
        if raw[0] == 0x47 and raw[1] == 0x49:
            return 'image/gif'
        if raw[0] == 0x52 and raw[1] == 0x49:
            return 'image/webp'
    except:
        pass
    return None


def parse_excel_content(b64_data: str, filename: str = "file.xlsx") -> str:
    if not EXCEL_SUPPORT:
        return f"[Fichier Excel: {filename} - openpyxl non installe]"
    try:
        excel_bytes = base64.b64decode(b64_data)
        excel_file = io.BytesIO(excel_bytes)
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        result_parts = [f"=== CONTENU FICHIER EXCEL: {filename} ==="]
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            result_parts.append(f"\n--- Feuille: {sheet_name} ---")
            rows_content = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    rows_content.append(row_str)
            if rows_content:
                result_parts.extend(rows_content[:100])
            else:
                result_parts.append("(Feuille vide)")
        wb.close()
        return "\n".join(result_parts)
    except Exception as e:
        return f"[Erreur lecture Excel {filename}: {str(e)}]"


def build_prompt(history_text, date_str, body_text, photos_count, excel_content):
    return f"""Coach ELITE transformation physique. Analyse ce bilan HONNETEMENT et reponds en JSON.

STYLE: Direct, tutoiement, JAMAIS d'asterisques (*), HONNETE sur le physique, pas de flatterie.

GUIDE MASSE GRASSE (STRICT):
FEMME: 18-20%=fit+abdos visibles | 24-28%=normale, PAS d'abdos, gras hanches/cuisses | 30-35%=surpoids | 35%+=obesite
HOMME: 10-12%=abdos decoupes | 15-18%=fit, abdos peu visibles | 20-25%=gras ventre | 28%+=surpoids
REGLE: Pas d'abdos visibles = MINIMUM 25% femme / 20% homme.

CONSEILS ADAPTES AU PROFIL:
- SURPOIDS: JAMAIS de snacks/barres! Structure et discipline sur les repas.
- SEC en seche: collation de secours OK.

=== HISTORIQUE COMPLET DEPUIS JOUR 1 ===
{history_text[:6000] if history_text else "Premier contact"}

=== BILAN ACTUEL ({date_str}) ===
{body_text[:5000]}

DONNEES: {photos_count} photos, Excel: {bool(excel_content)}
{excel_content[:2000] if excel_content else ""}

EXTRACTION (LIS TOUT L'HISTORIQUE!):
1. POIDS DE DEPART: Premier email client
2. POIDS ACTUEL: Bilan actuel
3. PROGRAMME: Dans les reponses COACH de l'historique
4. NE JAMAIS INVENTER!

CONSEILS ULTRA-SPECIFIQUES (JAMAIS DE VAGUE!):
Tu es un coach EXPERT avec 11 certifications. JAMAIS de conseils generiques!

INTERDIT: "ajoute du volume", "resserre la diet", "fais plus de cardio"
OBLIGATOIRE:
- EXERCICES PRECIS: "4x10 developpe incline + 3x12 ecarte poulie"
- MACROS EXACTES: "200P / 180G / 60L = 2060kcal"
- DUREE/FREQUENCE: "35min LISS, 4x/semaine"

EMAIL HTML OBLIGATOIRE - Style ACHZOD Premium:
Utilise ce template EXACT avec les vraies donnees:

<div style="font-family:'Helvetica Neue',Arial,sans-serif;max-width:600px;margin:0 auto;background:#0f0f1e;padding:0;">
  <!-- HEADER ACHZOD -->
  <div style="background:linear-gradient(135deg,#7c3aed 0%,#a855f7 50%,#ec4899 100%);padding:32px 24px;text-align:center;border-radius:0 0 24px 24px;">
    <h1 style="color:white;margin:0;font-size:28px;font-weight:800;letter-spacing:-0.5px;">ACHZOD COACHING</h1>
    <p style="color:rgba(255,255,255,0.9);margin:8px 0 0;font-size:14px;">Ton analyse personnalisee</p>
  </div>

  <!-- SALUTATION -->
  <div style="padding:32px 24px 16px;color:#e0e0e0;">
    <p style="font-size:18px;margin:0;">Salut <strong style="color:#a855f7;">[PRENOM]</strong> 💪</p>
    <p style="color:#9ca3af;margin:12px 0 0;line-height:1.6;">[INTRO PERSONNALISEE]</p>
  </div>

  <!-- BLOC EVOLUTION - GRADIENT -->
  <div style="margin:24px;background:linear-gradient(135deg,#1e1b4b 0%,#312e81 100%);border-radius:20px;padding:24px;border:1px solid rgba(139,92,246,0.3);">
    <h2 style="color:white;margin:0 0 20px;font-size:18px;text-align:center;">📈 Ton Evolution depuis le Jour 1</h2>
    <div style="display:flex;flex-wrap:wrap;gap:12px;justify-content:center;">
      <div style="background:rgba(255,255,255,0.1);border-radius:16px;padding:20px;text-align:center;min-width:130px;flex:1;">
        <div style="color:#a78bfa;font-size:13px;text-transform:uppercase;letter-spacing:1px;">Poids</div>
        <div style="color:white;font-size:22px;font-weight:700;margin:8px 0;">[XX]kg → [XX]kg</div>
        <div style="color:#4ade80;font-size:18px;font-weight:600;">[+/-X]kg</div>
      </div>
      <div style="background:rgba(255,255,255,0.1);border-radius:16px;padding:20px;text-align:center;min-width:130px;flex:1;">
        <div style="color:#a78bfa;font-size:13px;text-transform:uppercase;letter-spacing:1px;">Masse Grasse</div>
        <div style="color:white;font-size:22px;font-weight:700;margin:8px 0;">[XX]%</div>
        <div style="color:#fbbf24;font-size:14px;">[Evolution]</div>
      </div>
    </div>
  </div>

  <!-- ANALYSE PHOTOS si applicable -->
  <div style="margin:24px;background:#1a1a2e;border-radius:16px;padding:24px;border-left:4px solid #a855f7;">
    <h3 style="color:#a855f7;margin:0 0 16px;font-size:16px;">📸 Analyse Visuelle</h3>
    <p style="color:#d1d5db;margin:0;line-height:1.7;">[ANALYSE DES PHOTOS - evolution visible, zones ameliorees, zones a travailler]</p>
  </div>

  <!-- CE QUI VA BIEN -->
  <div style="margin:24px;background:linear-gradient(135deg,rgba(34,197,94,0.1),rgba(34,197,94,0.05));border-radius:16px;padding:24px;border-left:4px solid #22c55e;">
    <h3 style="color:#22c55e;margin:0 0 16px;font-size:16px;">✅ Ce qui cartonne</h3>
    <ul style="color:#d1d5db;margin:0;padding-left:20px;line-height:1.8;">
      <li>[Point positif 1]</li>
      <li>[Point positif 2]</li>
    </ul>
  </div>

  <!-- PLAN D'ACTION -->
  <div style="margin:24px;background:linear-gradient(135deg,rgba(251,191,36,0.1),rgba(251,191,36,0.05));border-radius:16px;padding:24px;border-left:4px solid #fbbf24;">
    <h3 style="color:#fbbf24;margin:0 0 16px;font-size:16px;">🎯 Plan d'Action Cette Semaine</h3>
    <div style="color:#d1d5db;line-height:1.8;">
      <p style="margin:0 0 12px;"><strong style="color:#a855f7;">Training:</strong> [EXERCICES PRECIS avec series/reps]</p>
      <p style="margin:0 0 12px;"><strong style="color:#a855f7;">Nutrition:</strong> [MACROS EXACTES ou ajustements]</p>
      <p style="margin:0;"><strong style="color:#a855f7;">Cardio:</strong> [DUREE et FREQUENCE precise]</p>
    </div>
  </div>

  <!-- FOOTER -->
  <div style="padding:32px 24px;text-align:center;border-top:1px solid rgba(255,255,255,0.1);">
    <p style="color:#9ca3af;margin:0 0 8px;font-size:14px;">On lache rien! 🔥</p>
    <p style="color:#a855f7;margin:0;font-size:20px;font-weight:700;">Achzod</p>
  </div>
</div>

JSON:
{{"resume":"analyse","analyse_photos":{{"masse_grasse_actuelle":"X%","evolution":"desc"}},"evolution_metriques":{{"poids":{{"jour1":"Xkg","actuel":"Xkg","diff":"-Xkg"}}}},"points_positifs":["x"],"points_ameliorer":[{{"probleme":"x","solution":"y"}}],"plan_action":{{"training":"x","nutrition":"x","cardio":"x"}},"draft_email":"HTML COMPLET"}}"""


def call_claude(prompt: str, images: list) -> Dict:
    """Appel Claude Sonnet 4.5 via Anthropic API"""
    try:
        content = [{"type": "text", "text": prompt}]
        for img_data, img_type in images:
            content.append({
                "type": "image",
                "source": {
                    "type": "base64",
                    "media_type": img_type,
                    "data": img_data
                }
            })

        response = anthropic_client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=8000,
            messages=[{"role": "user", "content": content}]
        )
        return {"success": True, "text": response.content[0].text, "model": "Claude-Sonnet-4.5"}
    except Exception as e:
        return {"success": False, "error": str(e), "model": "Claude-Sonnet-4.5"}


def call_gemini(prompt: str, images: list) -> Dict:
    """Appel Gemini via REST API avec fallback automatique"""

    # Liste des modeles a essayer (du plus recent au plus stable)
    models_to_try = [
        "gemini-2.0-flash-exp",
        "gemini-1.5-pro-latest",
        "gemini-1.5-flash"
    ]

    # Build parts
    parts = [{"text": prompt}]
    for img_data, img_type in images:
        parts.append({
            "inline_data": {
                "mime_type": img_type,
                "data": img_data
            }
        })

    # Config pour precision (pas d'hallucination)
    payload = {
        "contents": [{"parts": parts}],
        "generationConfig": {
            "temperature": 0.2,
            "topP": 0.95,
            "topK": 40,
            "maxOutputTokens": 8192
        }
    }

    last_error = ""
    for model_name in models_to_try:
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent?key={GOOGLE_API_KEY}"
            response = requests.post(url, json=payload, timeout=120)
            data = response.json()

            if "candidates" in data and data["candidates"]:
                text = data["candidates"][0]["content"]["parts"][0]["text"]
                return {"success": True, "text": text, "model": f"Gemini ({model_name})"}
            else:
                last_error = data.get("error", {}).get("message", "No candidates")
                continue  # Try next model
        except Exception as e:
            last_error = str(e)
            continue

    return {"success": False, "error": last_error, "model": "Gemini"}


def parse_json_response(response_text: str) -> Dict:
    """Parse JSON depuis la reponse IA"""
    try:
        text = response_text.strip()
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            for p in text.split("```"):
                if "{" in p and "draft_email" in p:
                    text = p
                    break

        start = text.find("{")
        end = text.rfind("}")
        if start != -1 and end != -1:
            json_str = text[start:end+1]
            return json.loads(json_str)
    except:
        pass

    # Fallback: extraire les champs manuellement
    analysis = {"resume": "", "draft_email": "", "points_positifs": [], "points_ameliorer": []}

    # Extraire draft_email
    draft_match = re.search(r'"draft_email"\s*:\s*"((?:[^"\\]|\\.)*)"', response_text, re.DOTALL)
    if draft_match:
        analysis["draft_email"] = draft_match.group(1).replace('\\n', '\n').replace('\\"', '"')

    return analysis


def analyze_coaching_bilan(current_email, conversation_history, client_name=""):
    """Analyse avec GPT-5.2 ET Gemini 3 Pro - retourne les 2 resultats"""

    history_text = _build_history_context(conversation_history)
    attachments = current_email.get("attachments", [])
    photos = [att for att in attachments if att.get("content_type", "").startswith("image/")]
    excels = [att for att in attachments if any(ext in att.get("filename", "").lower() for ext in ['.xlsx', '.xls'])]

    # Parse Excel
    excel_content = ""
    for excel_att in excels:
        if excel_att.get("data"):
            excel_content += "\n\n" + parse_excel_content(excel_att["data"], excel_att.get("filename", "fichier.xlsx"))

    date_str = current_email["date"].strftime("%d/%m/%Y %H:%M") if current_email.get("date") else "N/A"
    body_text = current_email.get("body", "") or ""

    # Parse Google Docs links in body
    gdoc_content = parse_all_google_docs(body_text)
    if gdoc_content:
        print(f"[ANALYZE] Google Docs content added ({len(gdoc_content)} chars)")

    # Build prompt with Google Docs content
    prompt = build_prompt(history_text, date_str, body_text, len(photos), excel_content + gdoc_content)

    # Prepare images
    images = []
    VALID_TYPES = ["image/jpeg", "image/png", "image/gif", "image/webp"]
    for att in photos[:5]:
        try:
            real_type = detect_image_type(att["data"])
            if real_type and real_type in VALID_TYPES:
                compressed, final_type = compress_image_if_needed(att["data"], real_type)
                images.append((compressed, final_type))
        except:
            pass

    print(f"[ANALYZE] Running Claude Sonnet 4.5 and Gemini in parallel with {len(images)} images...")

    # Appels en parallele
    results = {"claude": None, "gemini": None}

    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = {
            executor.submit(call_claude, prompt, images): "claude",
            executor.submit(call_gemini, prompt, images): "gemini"
        }
        for future in as_completed(futures):
            ai_name = futures[future]
            try:
                results[ai_name] = future.result()
            except Exception as e:
                results[ai_name] = {"success": False, "error": str(e), "model": ai_name}

    # Parser les resultats
    analyses = {}

    for ai_name, result in results.items():
        if result and result.get("success"):
            parsed = parse_json_response(result["text"])
            parsed["_model"] = result["model"]
            parsed["_raw"] = result["text"][:500]
            analyses[ai_name] = parsed
        else:
            analyses[ai_name] = {"error": result.get("error", "Unknown error"), "_model": result.get("model", ai_name)}

    # Retourner les 2 analyses
    return {
        "success": True,
        "gpt4": analyses.get("claude", {}),  # Claude remplace GPT
        "gemini": analyses.get("gemini", {}),
        "photos_analyzed": len(images)
    }


def _build_history_context(history):
    if not history:
        return ""
    parts = [f"[{len(history)} emails precedents]"]
    for i, e in enumerate(history[-15:], 1):
        if not e or not isinstance(e, dict):
            continue
        direction = "CLIENT" if e.get("direction") == "received" else "COACH (toi)"
        dt = e.get("date").strftime("%d/%m/%Y") if e.get("date") else "?"
        body = e.get('body', '') or ''
        body_preview = body[:1500] + "..." if len(body) > 1500 else body
        parts.append(f"\n--- Email {i} - {direction} ({dt}) ---\n{body_preview}")
    return "\n".join(parts)


def regenerate_email_draft(analysis, instructions, current_draft):
    """Regenere le draft avec Claude Sonnet 4.5"""
    prompt = f"""Tu es Achzod, coach expert. JAMAIS d'asterisques.
Analyse: {json.dumps(analysis, ensure_ascii=False)[:3000]}
Draft actuel: {current_draft}
Instructions: {instructions}
Reecris email 250-400 mots MAXIMUM, sans asterisques, style direct expert tutoiement."""

    try:
        r = anthropic_client.messages.create(
            model="claude-sonnet-4-5-20250929",
            max_tokens=2000,
            messages=[{"role": "user", "content": prompt}]
        )
        return r.content[0].text
    except Exception as e:
        return f"Erreur: {e}"
